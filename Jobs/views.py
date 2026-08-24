import openpyxl
from django.http import HttpResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.models import User, Group
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from .models import CandidateProfile, RecruiterProfile, Job, Application, SavedJob
from django.core.mail import send_mail
from django.core.paginator import Paginator
from .models import SavedJob
from django.shortcuts import redirect
from django.contrib import messages


def is_recruiter(user):
    return user.groups.filter(name='Recruiter').exists()


def is_candidate(user):
    return user.groups.filter(name='Candidate').exists()

def is_recruiter(user):
    return user.groups.filter(name='Recruiter').exists()

def is_candidate(user):
    return user.groups.filter(name='Candidate').exists()


def home(request):
    return render(request, 'home.html')


def candidate_register(request):
    if request.method == 'POST':
        username = request.POST['username']
        email = request.POST['email']
        password = request.POST['password']
        phone = request.POST['phone']
        qualification = request.POST['qualification']
        skills = request.POST['skills']
        resume = request.FILES.get('resume')

        user = User.objects.create_user(
            username=username,
            email=email,
            password=password
        )

        group, created = Group.objects.get_or_create(name='Candidate')
        user.groups.add(group)

        CandidateProfile.objects.create(
            user=user,
            phone=phone,
            qualification=qualification,
            skills=skills,
            resume=resume
        )

        return redirect('login')

    return render(request, 'candidate_register.html')


def recruiter_register(request):
    if request.method == 'POST':
        username = request.POST['username']
        email = request.POST['email']
        password = request.POST['password']
        company_name = request.POST['company_name']
        phone = request.POST['phone']
        company_location = request.POST['company_location']

        user = User.objects.create_user(
            username=username,
            email=email,
            password=password
        )

        group, created = Group.objects.get_or_create(name='Recruiter')
        user.groups.add(group)

        RecruiterProfile.objects.create(
            user=user,
            company_name=company_name,
            phone=phone,
            company_location=company_location
        )

        return redirect('login')

    return render(request, 'recruiter_register.html')


def user_login(request):
    if request.method == 'POST':
        username = request.POST['username']
        password = request.POST['password']

        user = authenticate(username=username, password=password)

        if user is not None:
            login(request, user)
            return redirect('dashboard')
        else:
            return render(request, 'login.html', {'error': 'Invalid username or password'})

    return render(request, 'login.html')


def user_logout(request):
    logout(request)
    return redirect('home')
@login_required
def dashboard(request):
    jobs = Job.objects.filter(recruiter=request.user)

    recent_applications = Application.objects.filter(
        job__recruiter=request.user
    ).order_by('-applied_date')[:5]

    saved_job_ids = SavedJob.objects.filter(
        candidate=request.user
        ).values_list('job_id', flat=True)

    context = {
        'total_jobs': jobs.count(),
        'my_posted_jobs': jobs.count(),
        'open_jobs': jobs.filter(status='Open').count(),
        'closed_jobs': jobs.filter(status='Closed').count(),
        'total_applications': Application.objects.filter(job__recruiter=request.user).count(),
        'recent_jobs': jobs.order_by('-posted_date')[:5],
        'recent_applications': recent_applications,
        'saved_jobs_count': SavedJob.objects.count(),
        'saved_job_ids': saved_job_ids,
    }

    return render(request, 'dashboard.html', context)

@login_required
def job_list(request):
    jobs = Job.objects.all().order_by('-posted_date')

    search = request.GET.get('search')
    location = request.GET.get('location')
    skills = request.GET.get('skills')

    if search:
        jobs = jobs.filter(title__icontains=search)

    if location:
        jobs = jobs.filter(location__icontains=location)

    if skills:
        jobs = jobs.filter(skills_required__icontains=skills)

    applied_job_ids = Application.objects.filter(
        candidate=request.user
    ).values_list('job_id', flat=True)

    paginator = Paginator(jobs, 5)
    page_number = request.GET.get('page')
    jobs = paginator.get_page(page_number)

    return render(request, 'job_list.html', {
        'jobs': jobs,
        'applied_job_ids': applied_job_ids,
    })

@login_required
def post_job(request):
    if not is_recruiter(request.user):
        return redirect('dashboard')

    if request.method == 'POST':
        title = request.POST['title']
        company = request.POST['company']
        location = request.POST['location']
        salary = request.POST['salary']
        skills_required = request.POST['skills_required']
        description = request.POST['description']
        company_logo = request.FILES.get('company_logo')

        Job.objects.create(
            recruiter=request.user,
            title=title,
            company=company,
            location=location,
            salary=salary,
            skills_required=skills_required,
            description=description,
            company_logo=company_logo,
        )

        messages.success(request, "Job posted successfully!")
        return redirect('job_list')

    return render(request, 'post_job.html')

@login_required
def apply_job(request, job_id):
    job = get_object_or_404(Job, id=job_id)

    if job.status == "Closed":
        return redirect('job_list')

    already_applied = Application.objects.filter(
        candidate=request.user,
        job=job
    ).exists()

    


    if not already_applied:
        Application.objects.create(
            candidate=request.user,
            job=job
        )

        send_mail(
            subject='New Job Application',
            message=f'{request.user.username} applied for {job.title}',
            from_email='admin@jobportal.com',
            recipient_list=[job.recruiter.email],
            fail_silently=True,
    )

        messages.success(request, "Job applied successfully!")

    else:
        messages.warning(
            request,
            "You have already applied for this job."
        )

    return redirect('my_applications')


@login_required
def my_applications(request):
    applications = Application.objects.filter(candidate=request.user)
    return render(request, 'my_applications.html', {'applications': applications})

@login_required
def recruiter_dashboard(request):
    jobs = Job.objects.filter(recruiter=request.user)

    total_jobs = jobs.count()
    open_jobs = jobs.filter(status="Open").count()
    closed_jobs = jobs.filter(status="Closed").count()

    total_applications = Application.objects.filter(job__recruiter=request.user).count()

    recent_applications = Application.objects.filter(
        job__recruiter=request.user
    ).order_by("-applied_date")[:5]

    context = {
        "total_jobs": total_jobs,
        "open_jobs": open_jobs,
        "closed_jobs": closed_jobs,
        "total_applications": total_applications,
        "recent_applications": recent_applications,
    }

    return render(request, "recruiter_dashboard.html", context)

@login_required
def profile(request):
    candidate_profile = None
    recruiter_profile = None

    if CandidateProfile.objects.filter(user=request.user).exists():
        candidate_profile = CandidateProfile.objects.get(user=request.user)

    if RecruiterProfile.objects.filter(user=request.user).exists():
        recruiter_profile = RecruiterProfile.objects.get(user=request.user)

    return render(request, 'profile.html', {
        'candidate_profile': candidate_profile,
        'recruiter_profile': recruiter_profile,
    })

@login_required
def edit_profile(request):
    candidate_profile = None
    recruiter_profile = None

    if CandidateProfile.objects.filter(user=request.user).exists():
        candidate_profile = CandidateProfile.objects.get(user=request.user)

    if RecruiterProfile.objects.filter(user=request.user).exists():
        recruiter_profile = RecruiterProfile.objects.get(user=request.user)

    if request.method == 'POST':
        request.user.email = request.POST.get('email')
        request.user.save()

        if candidate_profile:
            candidate_profile.phone = request.POST.get('phone')
            candidate_profile.qualification = request.POST.get('qualification')
            candidate_profile.skills = request.POST.get('skills')

            if request.FILES.get('resume'):
                candidate_profile.resume = request.FILES.get('resume')

            if request.FILES.get('profile_photo'):
                candidate_profile.profile_photo = request.FILES.get('profile_photo')

            candidate_profile.save()

        if recruiter_profile:
            recruiter_profile.phone = request.POST.get('phone')
            recruiter_profile.company_name = request.POST.get('company_name')
            recruiter_profile.company_location = request.POST.get('company_location')

            if request.FILES.get('profile_photo'):
                recruiter_profile.profile_photo = request.FILES.get('profile_photo')

            recruiter_profile.save()

        return redirect('profile')

    return render(request, 'edit_profile.html', {
        'candidate_profile': candidate_profile,
        'recruiter_profile': recruiter_profile,
    })

@login_required
def job_applicants(request, job_id):
    if not is_recruiter(request.user):
        return redirect('dashboard')
    
    job = get_object_or_404(Job, id=job_id, recruiter=request.user)

    applications = Application.objects.filter(job=job)

    status = request.GET.get('status')

    if status:
        applications = applications.filter(status=status)

    return render(request, 'job_applicants.html', {
        'job': job,
        'applications': applications,
        'selected_status': status,
    })

@login_required
def edit_job(request, job_id):
    if not is_recruiter(request.user):
        return redirect('dashboard')
    
    job = get_object_or_404(Job, id=job_id, recruiter=request.user)

    if request.method == 'POST':
        job.title = request.POST.get('title')
        job.company = request.POST.get('company')
        job.location = request.POST.get('location')
        job.salary = request.POST.get('salary')
        job.skills_required = request.POST.get('skills_required')
        job.description = request.POST.get('description')
        job.status = request.POST.get('status')
        job.save()

        return redirect('my_posted_jobs')

    return render(request, 'edit_job.html', {'job': job})


@login_required
def delete_job(request, job_id):
    if not is_recruiter(request.user):
        return redirect('dashboard')
    
    job = get_object_or_404(Job, id=job_id, recruiter=request.user)

    if request.method == 'POST':
        job.delete()
        return redirect('my_posted_jobs')

    return render(request, 'delete_job.html', {'job': job})

@login_required
def save_job(request, job_id):
    job = get_object_or_404(Job, id=job_id)

    SavedJob.objects.get_or_create(
        candidate=request.user,
        job=job
    )

    messages.success(request, "Job saved successfully!")

    return redirect('job_list')


@login_required
def saved_jobs(request):
    saved_jobs = SavedJob.objects.filter(candidate=request.user)

    return render(request, 'saved_jobs.html', {
        'saved_jobs': saved_jobs
    })

@login_required
def update_application_status(request, application_id):
    if not is_recruiter(request.user):
        return redirect('dashboard')
    
    application = get_object_or_404(
        Application,
        id=application_id,
        job__recruiter=request.user
    )

    if request.method == 'POST':
        new_status = request.POST.get('status')
        application.status = new_status
        application.save()

        if new_status == 'Selected':
            send_mail(
                'Application Status Update',
                f'Congratulations {application.candidate.username}, you have been selected for {application.job.title}.',
                'admin@jobportal.com',
                [application.candidate.email],
                fail_silently=True,
            )

    return redirect('job_applicants', job_id=application.job.id)

@login_required
def candidate_profile(request, user_id):
    candidate = get_object_or_404(User, id=user_id)

    profile = get_object_or_404(
        CandidateProfile,
        user=candidate
    )

    return render(request, 'candidate_profile.html', {
        'candidate': candidate,
        'profile': profile,
    })

@login_required
def my_posted_jobs(request):
    if not is_recruiter(request.user):
        return redirect('dashboard')

    jobs = Job.objects.filter(recruiter=request.user).order_by('-posted_date')
    return render(request, 'my_posted_jobs.html', {'jobs': jobs})


@login_required
def toggle_save_job(request, job_id):
    job = get_object_or_404(Job, id=job_id)

    saved_job = SavedJob.objects.filter(
        candidate=request.user,
        job=job
    )

    if saved_job.exists():
        saved_job.delete()
        messages.warning(request, "Job removed from saved jobs.")
    else:
        SavedJob.objects.create(
            candidate=request.user,
            job=job
        )
        messages.success(request, "Job saved successfully!")

    return redirect('job_list')

@login_required
def export_applicants_excel(request, job_id):
    if not is_recruiter(request.user):
        return redirect('dashboard')

    job = get_object_or_404(Job, id=job_id, recruiter=request.user)

    applications = Application.objects.filter(job=job)

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Applicants"

    headers = [
        "Candidate Name",
        "Email",
        "Job Title",
        "Status",
        "Applied Date"
    ]

    for col_num, header in enumerate(headers, 1):
        sheet.cell(row=1, column=col_num).value = header

    row_num = 2

    for app in applications:
        sheet.cell(row=row_num, column=1).value = app.candidate.username
        sheet.cell(row=row_num, column=2).value = app.candidate.email
        sheet.cell(row=row_num, column=3).value = app.job.title
        sheet.cell(row=row_num, column=4).value = app.status
        sheet.cell(row=row_num, column=5).value = str(app.applied_date)

        row_num += 1

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

    response['Content-Disposition'] = (
        f'attachment; filename=Applicants_{job.id}.xlsx'
    )

    workbook.save(response)

    return response
